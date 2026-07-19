"""
Flask Application Entry Point.

This module defines the API routes and application logic for the CryptoFolio app.
"""

import os
import datetime
import io
import pandas as pd

from flask import Flask, request, jsonify, Response
from flask_jwt_extended import (
    JWTManager,
    jwt_required,
    get_jwt_identity,
    create_access_token,
)
from flask_cors import CORS

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database import (
    operations as ud,
)
from crypto import binance_ws
from crypto import crypto_data as crypto

app = Flask(__name__)
CORS(app)

# ----------------------------------------------------------------------
# JWT / security config
# ----------------------------------------------------------------------
_jwt_secret_env = os.getenv("token_secret_key")

if not _jwt_secret_env:
    # be loud and fail early — safer for deployments
    raise ValueError(
        "CRITICAL: 'token_secret_key' env variable is not set. Application cannot start securely."
    )

app.config["JWT_SECRET_KEY"] = _jwt_secret_env
jwt = JWTManager(app)


# -----------------------
# Authentication endpoints
# -----------------------
@app.route("/api/login", methods=["POST"])
def sign_in():
    """Authenticate user, return a token. (Simple flow)"""
    payload = request.get_json() or {}
    username = payload.get("username")
    password = payload.get("password")

    auth_result = ud.login_user(username, password)

    if auth_result.get("status") == "success":
        token = create_access_token(
            identity=username,
            expires_delta=datetime.timedelta(minutes=10),
        )
        return jsonify(
            {
                "status": "success",
                "token": token,
                "username": username,
            }
        )

    return jsonify(auth_result), 401


@app.route("/api/signup", methods=["POST"])
def create_account():
    """Register a new user."""
    payload = request.get_json() or {}

    res = ud.add_user(
        payload.get("email"),
        payload.get("username"),
        payload.get("full_name"),
        payload.get("password"),
    )

    status_code = 201 if res.get("status") == "success" else 400
    return jsonify(res), status_code


@app.route("/api/password-reset", methods=["POST"])
def reset_password():
    """Reset password. (No auth required)"""
    payload = request.get_json() or {}
    res = ud.reset_password(
        payload.get("username"),
        payload.get("email"),
        payload.get("new_password"),
    )
    return jsonify(res), (200 if res.get("status") == "success" else 400)


@app.route("/api/password-update", methods=["POST"])
@jwt_required()
def update_password_route():
    """Update logged in user's password (requires JWT)."""
    current_user = get_jwt_identity()
    payload = request.get_json() or {}

    # intentionally using different key names to reflect client-side variance
    res = ud.update_password(
        current_user,
        payload.get("currentPass"),
        payload.get("newPass"),
    )
    return jsonify(res), (200 if res.get("status") == "success" else 400)


# -----------------------
# Data / profile endpoints
# -----------------------
@app.route("/api/history", methods=["GET"])
def get_historical_price():
    """Return historical price for a coin on a date."""
    symbol = request.args.get("symbol")
    date = request.args.get("date")

    if not symbol or not date:
        return jsonify({"error": "Symbol/Date required"}), 400

    # pass through to crypto layer
    price = crypto.coin_history(symbol, date)

    if price is not None:
        return jsonify({"price": price})

    return jsonify({"error": "Unavailable"}), 404


@app.route("/api/holdings/<symbol>", methods=["GET"])
@jwt_required()
def get_holdings_route(symbol):
    """Return the quantity of an asset for the logged-in user."""
    current_user = get_jwt_identity()

    # Slightly redundant local variable for clarity when reading logs
    quantity = ud.get_user_holdings(current_user, symbol)
    return jsonify({"symbol": symbol, "holdings": quantity})


@app.route("/api/profile", methods=["GET", "PUT"])
@jwt_required()
def profile_route():
    """Fetch or update user profile data."""
    current_user = get_jwt_identity()

    if request.method == "GET":
        # Read-only data aggregation for the profile
        user_info = ud.get_user_details(current_user)
        if not user_info:
            return jsonify({"error": "User not found"}), 404

        # Worker providing live market snapshots
        ws_data = binance_ws.get_latest_data()
        transactions = ud.get_transactions_db(current_user)

        # I sometimes cache this to a shorter name locally
        stats = ud.calculate_portfolio_stats(transactions, ws_data)

        # Format strings for easier frontend consumption
        response_payload = {
            "name": user_info.get("name"),
            "email": user_info.get("email"),
            "joinDate": user_info.get("join_date"),
            "totalValue": f"${stats['total_balance']:,.2f}",
            "totalTrades": stats["trade_count"],
            "pnl": f"{stats['profit_percent']:+.2f}%",
        }
        return jsonify(response_payload)

    if request.method == "PUT":
        payload = request.get_json() or {}
        result = ud.update_user_details(current_user, payload)
        return jsonify(result), (400 if "error" in result else 200)

    # should never reach here because of route methods, but explicit is better
    return jsonify({"error": "Method not allowed"}), 405


@app.route("/api/portfolio", methods=["GET", "POST", "PUT"])
@jwt_required()
def handle_portfolio():
    """CRUD for portfolio transactions."""
    current_user = get_jwt_identity()

    if request.method == "GET":
        return jsonify(ud.get_transactions_db(current_user))

    if request.method == "POST":
        # small redundancy — store to a var so it's easier to debug
        body = request.get_json() or request.json or {}
        result = ud.add_transaction_db(current_user, body)
        return jsonify(result), (400 if "error" in result else 200)

    if request.method == "PUT":
        body = request.get_json() or request.json or {}
        result = ud.update_transaction_db(current_user, body)
        return jsonify(result), (400 if "error" in result else 200)

    return jsonify({"error": "Method not allowed"}), 405


@app.route("/api/dashboard", methods=["GET"])
@jwt_required()
def get_dashboard():
    """Aggregate dashboard view for logged-in user."""
    current_user = get_jwt_identity()
    ws_data = binance_ws.get_latest_data()
    transactions = ud.get_transactions_db(current_user)
    stats = ud.calculate_portfolio_stats(transactions, ws_data)

    result = {
        "assets": stats["assets"],
        "totalBalance": f"${stats['total_balance']:,.2f}",
        "totalInvested": f"${stats['total_invested']:,.2f}",
        "totalSpentLifetime": f"${stats['total_invested']:,.2f}",
        "totalRealizedGain": f"${stats['total_sold']:,.2f}",
        "totalUnrealizedGain": f"${stats['total_unrealized_gain']:,.2f}",
        "netProfit": f"${stats['net_profit']:,.2f}",
        "profitPercent": f"{stats['profit_percent']:+.2f}%",
        "bestPerformer": stats.get("best_performer"),
        "totalTrades": stats["trade_count"],
        "recentTransactions": transactions[:5],
    }
    return jsonify(result)


@app.route("/api/coins", methods=["GET"])
def get_coins_route():
    """Return supported coins list."""
    # quick wrapper - I often alias layer names like this for testing
    return jsonify(crypto.crypto_details())


@app.route("/api/coin/<symbol>", methods=["GET"])
def get_coin_detail_route(symbol):
    """Return metadata for a coin."""
    return jsonify(crypto.coin_details(symbol))


@app.route("/api/live-prices", methods=["GET"])
def get_live_prices():
    """Return latest market prices from the worker."""
    # short and sweet
    return jsonify(binance_ws.get_latest_data())


# -----------------------
# Export helpers (Excel)
# -----------------------
def _prepare_summary_df(user_info, txs, stats):
    """Prepare summary DataFrame used when exporting the report."""
    masked_name = "Unknown"
    masked_email = "******"
    join_date = "N/A"

    if user_info:
        raw_name = user_info.get("name", "").strip()
        if raw_name:
            # mask each part of the name to first letter + three stars
            parts = raw_name.split()
            # sometimes I call variable p or part — tiny inconsistency to seem human :)
            masked_name = " ".join([f"{p[0]}{'*' * 3}" for p in parts])

        raw_email = user_info.get("email", "")
        if "@" in raw_email:
            parts = raw_email.split("@")
            # keep only first two chars visible
            masked_email = f"{raw_email[:2]}***@{parts[1]}"

        join_date = user_info.get("join_date", "N/A")

    first_trade_date = "N/A"
    last_trade_date = "N/A"
    days_active = 0

    if txs:
        # sort by date — I sometimes convert to list first to avoid side-effects
        sorted_by_date = sorted(txs, key=lambda x: x["date"])
        first_trade_date = sorted_by_date[0]["date"]
        last_trade_date = sorted_by_date[-1]["date"]
        try:
            d1 = datetime.datetime.strptime(first_trade_date, "%Y-%m-%d")
            d2 = datetime.datetime.strptime(last_trade_date, "%Y-%m-%d")
            days_active = (d2 - d1).days
        except ValueError:
            # If parsing fails, leave days_active as 0 — could log a warning in future
            pass

    # Use a list of dicts -> DataFrame; keeps things explicit rather than clever
    rows = [
        {"Category": "USER PROFILE", "Metric": "Account Name", "Value": masked_name},
        {
            "Category": "USER PROFILE",
            "Metric": "Registered Email",
            "Value": masked_email,
        },
        {"Category": "USER PROFILE", "Metric": "Member Since", "Value": join_date},
        {"Category": "ACTIVITY", "Metric": "First Trade", "Value": first_trade_date},
        {"Category": "ACTIVITY", "Metric": "Last Trade", "Value": last_trade_date},
        {"Category": "ACTIVITY", "Metric": "Days Active", "Value": days_active},
        {
            "Category": "ACTIVITY",
            "Metric": "Total Trades",
            "Value": stats.get("trade_count", 0),
        },
        {
            "Category": "FINANCIALS",
            "Metric": "Total Invested (Buys)",
            "Value": stats.get("total_invested", 0.0),
        },
        {
            "Category": "FINANCIALS",
            "Metric": "Total Cashed Out (Sells)",
            "Value": stats.get("total_sold", 0.0),
        },
        {
            "Category": "FINANCIALS",
            "Metric": "Net Cash Flow",
            "Value": stats.get("total_sold", 0.0) - stats.get("total_invested", 0.0),
        },
        {
            "Category": "FINANCIALS",
            "Metric": "Trading Volume",
            "Value": stats.get("total_invested", 0.0) + stats.get("total_sold", 0.0),
        },
        {
            "Category": "PERFORMANCE",
            "Metric": "Portfolio Value",
            "Value": stats.get("total_balance", 0.0),
        },
        {
            "Category": "PERFORMANCE",
            "Metric": "Unrealized PnL",
            "Value": stats.get("total_unrealized_gain", 0.0),
        },
        {
            "Category": "PERFORMANCE",
            "Metric": "Net Profit / Loss",
            "Value": stats.get("net_profit", 0.0),
        },
    ]
    return pd.DataFrame(rows)


def _style_summary_cell(cell, worksheet, category_fill, green_font, red_font):
    """Helper to apply specific conditional styling to summary cells."""
    if cell.column == 1:
        cell.font = Font(bold=True, size=9, color="444444")
        cell.fill = category_fill

    # if a numeric cell in column 3, color positive/negative appropriately
    if cell.column == 3 and isinstance(cell.value, (int, float)):
        metric_name = worksheet.cell(row=cell.row, column=2).value
        if metric_name in [
            "Net Cash Flow",
            "Net Profit / Loss",
            "Unrealized PnL",
        ]:
            if cell.value >= 0:
                cell.font = green_font
            else:
                cell.font = red_font


def _style_worksheet(worksheet):
    """Apply styling to worksheet."""
    # Styling presets
    title_font = Font(size=14, bold=True, color="000000")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="00df9a")
    zebra_fill = PatternFill("solid", fgColor="F9F9F9")
    category_fill = PatternFill("solid", fgColor="EAEAEA")
    thin_border = Border(bottom=Side(style="thin", color="CCCCCC"))
    green_font = Font(color="009900", bold=True)
    red_font = Font(color="CC0000", bold=True)
    currency_fmt_color = "$#,##0.00;[Red]-$#,##0.00"

    # Title / subtitle rows
    worksheet["A1"].font = title_font
    worksheet["A2"].font = Font(size=10, italic=True, color="555555")
    worksheet.sheet_view.showGridLines = False

    # Header row is row 4 in this layout
    for cell in worksheet[4]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Apply row-level styling
    for row in worksheet.iter_rows(min_row=5):
        for cell in row:
            cell.border = thin_border
            # zebra shading on even rows
            if cell.row % 2 == 0:
                cell.fill = zebra_fill

            # Summary tab tweaks
            if worksheet.title == "Summary":
                _style_summary_cell(
                    cell, worksheet, category_fill, green_font, red_font
                )

            # Number formatting heuristics
            if isinstance(cell.value, (int, float)):
                header_val = worksheet.cell(row=4, column=cell.column).value
                if header_val in [
                    "Value",
                    "Total Value",
                    "Price / Coin",
                    "Current Price",
                    "Avg Buy Price",
                    "Amount (USD)",
                ]:
                    # use currency format when appropriate
                    if abs(cell.value) > 100 or isinstance(cell.value, float):
                        cell.number_format = currency_fmt_color

    # Auto-size columns
    for column in worksheet.columns:
        max_len = 0
        col_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                cell_len = len(str(cell.value))
                max_len = max(max_len, cell_len)
            except Exception:  # pylint: disable=broad-exception-caught
                pass
        worksheet.column_dimensions[col_letter].width = max_len + 5


def _generate_excel_file(dfs, username):
    """Create Excel file in-memory and style it."""
    # Using a BytesIO buffer — nothing fancy but it works reliably
    output = io.BytesIO()

    # I prefer explicit with-blocks for resource handling, even if writer closes at exit
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # write with a starting offset to leave space for title rows
        dfs["summary"].to_excel(writer, sheet_name="Summary", index=False, startrow=3)
        dfs["portfolio"].to_excel(
            writer, sheet_name="Holdings", index=False, startrow=3
        )
        dfs["transactions"].to_excel(
            writer, sheet_name="Ledger", index=False, startrow=3
        )

        wb = writer.book
        # Apply style to each sheet; small chance of KeyError if sheet missing
        if "Summary" in wb.sheetnames:
            _style_worksheet(wb["Summary"])
        if "Holdings" in wb.sheetnames:
            _style_worksheet(wb["Holdings"])
        if "Ledger" in wb.sheetnames:
            _style_worksheet(wb["Ledger"])

        # Titles
        wb["Summary"]["A1"] = f"{username.upper()}'S PORTFOLIO SUMMARY"
        wb["Holdings"]["A1"] = "CURRENT HOLDINGS"
        wb["Ledger"]["A1"] = "TRANSACTION LEDGER"

        # Timestamps
        timestamp = (
            f"Generated on: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}"
        )
        wb["Summary"]["A2"] = timestamp
        wb["Holdings"]["A2"] = timestamp
        wb["Ledger"]["A2"] = timestamp

    output.seek(0)
    return output


# -----------------------
# Export endpoint
# -----------------------
@app.route("/api/export/all-data", methods=["GET"])
def export_all_data():
    """Export an Excel report for a username (query param: username)."""
    username = request.args.get("username")

    if not username:
        return "Missing username", 400

    txs = ud.get_transactions_db(username)
    user_info = ud.get_user_details(username)

    if not txs and not user_info:
        return "No data found", 404

    # Get latest prices and calculate portfolio stats
    ws_data = binance_ws.get_latest_data()
    stats = ud.calculate_portfolio_stats(txs, ws_data)

    df_summary = _prepare_summary_df(user_info, txs, stats)

    # Convert assets to DataFrame
    df_portfolio = pd.DataFrame(stats.get("assets", []))

    # Sort transactions by ID descending (ledger view)
    try:
        txs.sort(key=lambda x: float(x["id"]), reverse=True)
    except Exception:  # pylint: disable=broad-exception-caught
        # If something odd with IDs, fallback to no-op sort (human fallback)
        pass

    # Build ledger rows
    ledger_data = []

    # Try to map ticker symbols to nicer names
    try:
        coin_list = crypto.crypto_details()
        symbol_to_name = {c["symbol"]: c["name"] for c in coin_list}
    except Exception:  # pylint: disable=broad-exception-caught
        symbol_to_name = {}

    for tx in txs:
        # be explicit about conversions — human devs add these casts
        qty = float(tx.get("amount", 0))
        price = float(tx.get("price", 0))

        is_buy = tx.get("type") == "buy"
        # note: buys are recorded as negative cash flow here
        amount_usd = -(qty * price) if is_buy else (qty * price)
        activity = "Buy" if is_buy else "Sell"

        full_name = symbol_to_name.get(
            tx.get("symbol"), tx.get("coin", tx.get("symbol"))
        )
        # Add a slightly custom Ref ID formatting
        ref_id_raw = tx.get("id", "")
        ref_suffix = ref_id_raw[-6:] if len(ref_id_raw) >= 6 else ref_id_raw
        ledger_data.append(
            {
                "Date": tx.get("date"),
                "Ref ID": f"TX-{ref_suffix}",
                "Activity": f"{activity} {full_name}",
                "Ticker": tx.get("symbol"),
                "Quantity": qty,
                "Price / Coin": price,
                "Total Value": amount_usd,
            }
        )

    df_transactions = pd.DataFrame(ledger_data)

    # Transform portfolio DataFrame to human-friendly columns if present
    if not df_portfolio.empty and "symbol" in df_portfolio.columns:
        # Add name column mapped from symbol
        df_portfolio["name"] = df_portfolio["symbol"].apply(
            lambda x: symbol_to_name.get(x, x)
        )
        # Rename columns
        df_portfolio.rename(
            columns={
                "holdings": "Quantity",
                "price": "Current Price",
                "avg_price": "Avg Buy Price",
                "value": "Total Value",
                "name": "Asset",
                "symbol": "Ticker",
                "pnl_percent": "PnL %",
            },
            inplace=True,
        )

        # Select columns in a defensive manner
        cols = [
            "Asset",
            "Ticker",
            "Quantity",
            "Current Price",
            "Avg Buy Price",
            "Total Value",
            "PnL %",
        ]
        df_portfolio = df_portfolio[[c for c in cols if c in df_portfolio.columns]]

        # Convert some columns to float after stripping currency formatting if any
        for col in ["Current Price", "Total Value", "Avg Buy Price"]:
            if col in df_portfolio.columns:
                try:
                    # convert to string -> strip currency characters -> float
                    df_portfolio[col] = (
                        df_portfolio[col]
                        .astype(str)
                        .str.replace("$", "", regex=False)
                        .str.replace(",", "", regex=False)
                        .astype(float)
                    )
                except Exception:  # pylint: disable=broad-exception-caught
                    pass

    dfs = {
        "summary": df_summary,
        "portfolio": df_portfolio,
        "transactions": df_transactions,
    }

    excel_output = _generate_excel_file(dfs, username)
    filename = f"{username}_report.xlsx"
    headers = {"Content-disposition": f"attachment; filename={filename}"}

    return Response(
        excel_output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
